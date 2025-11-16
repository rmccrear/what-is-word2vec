import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import os

def create_excel_with_formulas(
    twitter_csv="top_1000_words_vectors_twitter.csv",
    wiki_csv="top_5000_words_vectors_wiki.csv",
    output_file="word_vectors_with_formulas.xlsx"
):
    """Create an Excel file with multiple word vector datasets and cosine similarity formulas."""
    
    # Read both CSV files
    print(f"Reading {twitter_csv}...")
    df_twitter = pd.read_csv(twitter_csv)
    
    print(f"Reading {wiki_csv}...")
    df_wiki = pd.read_csv(wiki_csv)
    
    # Create Excel writer
    print(f"Creating Excel file: {output_file}...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write Twitter data
        df_twitter.to_excel(writer, sheet_name='Word Vectors Twitter', index=False)
        
        # Write Wiki data
        df_wiki.to_excel(writer, sheet_name='Word Vectors Wiki', index=False)
        
        # Create a sheet for cosine similarity calculations
        similarity_df = pd.DataFrame({
            'Word 1': [],
            'Row Index 1': [],
            'Word 2': [],
            'Row Index 2': [],
            'Cosine Similarity': []
        })
        similarity_df.to_excel(writer, sheet_name='Cosine Similarity', index=False)
        
        # Create instructions sheet
        instructions = pd.DataFrame({
            'Instructions': [
                'COSINE SIMILARITY FORMULAS',
                '',
                'This workbook contains two datasets:',
                '- Word Vectors Twitter: GloVe Twitter 25-dimensional vectors',
                '- Word Vectors Wiki: GloVe Wiki-Gigaword 50-dimensional vectors',
                '',
                'Use the dropdown on the Search tab to select which dataset to use.',
                '',
                'To calculate cosine similarity between two vectors by row index:',
                '',
                'Method 1: Direct cell reference',
                '"=SUMPRODUCT(B2:Z2,B3:Z3)/(SQRT(SUMPRODUCT(B2:Z2,B2:Z2))*SQRT(SUMPRODUCT(B3:Z3,B3:Z3)))"',
                '',
                'Method 2: Using INDEX with row numbers',
                'If row indices are in cells A1 and B1:',
                '"=SUMPRODUCT(INDEX(B:Z,A1+1,0),INDEX(B:Z,B1+1,0))/(SQRT(SUMPRODUCT(INDEX(B:Z,A1+1,0),INDEX(B:Z,A1+1,0)))*SQRT(SUMPRODUCT(INDEX(B:Z,B1+1,0),INDEX(B:Z,B1+1,0))))"',
                '',
                'Note: +1 accounts for the header row. Row index 1 = actual row 2 in sheet.',
                '',
                'Result range: -1 to 1',
                '1 = most similar, 0 = orthogonal, -1 = most dissimilar'
            ]
        })
        instructions.to_excel(writer, sheet_name='Instructions', index=False)
        
        # Create a sheet for finding top matches (Search tab)
        top_matches_df = pd.DataFrame({
            'Rank': [str(i) for i in range(1, 11)],
            'Word': [''] * 10,
            'Similarity': [''] * 10
        })
        top_matches_df.to_excel(writer, sheet_name='Search', index=False)
    
    # Now add formulas and dropdown to Excel file
    print("Adding formulas and dropdown to Excel file...")
    wb = load_workbook(output_file)
    
    # Add formulas to Cosine Similarity sheet
    ws_similarity = wb['Cosine Similarity']
    
    # Define highlight color for input cells
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws_similarity.cell(row=1, column=1).value = "COSINE SIMILARITY CALCULATOR"
    ws_similarity.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_similarity.merge_cells('A1:E1')
    ws_similarity.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    ws_similarity.cell(row=2, column=1).value = "Instructions:"
    ws_similarity.cell(row=2, column=1).font = Font(bold=True)
    instructions_text = "Enter two words in columns A and C (starting from row 6). Select dataset using dropdown in B3. The sheet will automatically calculate cosine similarity. Similarity ranges from -1 (opposite) to 1 (identical)."
    ws_similarity.cell(row=2, column=2).value = instructions_text
    ws_similarity.merge_cells('B2:E2')
    ws_similarity.cell(row=2, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Add dataset selector dropdown
    ws_similarity.cell(row=3, column=1).value = "Dataset:"
    ws_similarity.cell(row=3, column=1).font = Font(bold=True)
    ws_similarity.cell(row=3, column=2).value = "Twitter"  # Default
    ws_similarity.cell(row=3, column=2).fill = input_fill  # Highlight input cell
    
    # Add data validation dropdown
    dv = DataValidation(type="list", formula1='"Twitter,Wiki"', allow_blank=False)
    ws_similarity.add_data_validation(dv)
    dv.add(ws_similarity.cell(row=3, column=2))
    
    headers = ['Word 1', 'Row Index 1', 'Word 2', 'Row Index 2', 'Cosine Similarity']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_similarity.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Example row with dynamic sheet reference
    # Using words that exist in both datasets: "people" and "back"
    ws_similarity.cell(row=6, column=1).value = "people"  # Example word 1 - exists in both datasets
    ws_similarity.cell(row=6, column=1).fill = input_fill  # Highlight input cell
    # Dynamic sheet name based on dropdown - need single quotes around sheet names with spaces
    # In Excel: INDIRECT("'Sheet Name'!A:A") - use double quotes for Python string to escape single quotes
    ws_similarity.cell(row=6, column=2).value = "=IF(A6=\"\",\"\",MATCH(A6,INDIRECT(\"'Word Vectors \"&B3&\"'!A:A\"),0)-1)"
    ws_similarity.cell(row=6, column=3).value = "back"  # Example word 2 - exists in both datasets
    ws_similarity.cell(row=6, column=3).fill = input_fill  # Highlight input cell
    ws_similarity.cell(row=6, column=4).value = "=IF(C6=\"\",\"\",MATCH(C6,INDIRECT(\"'Word Vectors \"&B3&\"'!A:A\"),0)-1)"
    
    # Cosine similarity formula with dynamic sheet reference
    # Twitter has 25 dims (B:Z), Wiki has 50 dims (B:AZ)
    # Note: Using double quotes for Python string allows proper escaping of single quotes
    formula = "=IF(OR(A6=\"\",C6=\"\",NOT(ISNUMBER(B6)),NOT(ISNUMBER(D6))),\"\",SUMPRODUCT(INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(B6+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(B6+1)),INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(D6+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(D6+1)))/(SQRT(SUMPRODUCT(INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(B6+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(B6+1)),INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(B6+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(B6+1))))*SQRT(SUMPRODUCT(INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(D6+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(D6+1)),INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(D6+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(D6+1))))))"
    ws_similarity.cell(row=6, column=5).value = formula
    
    # Template row
    ws_similarity.cell(row=7, column=1).value = ""
    ws_similarity.cell(row=7, column=1).fill = input_fill  # Highlight input cell
    ws_similarity.cell(row=7, column=2).value = "=IF(A7=\"\",\"\",IF(ISNA(MATCH(A7,INDIRECT(\"'Word Vectors \"&B3&\"'!A:A\"),0)),\"Word not found\",MATCH(A7,INDIRECT(\"'Word Vectors \"&B3&\"'!A:A\"),0)-1))"
    ws_similarity.cell(row=7, column=3).value = ""
    ws_similarity.cell(row=7, column=3).fill = input_fill  # Highlight input cell
    ws_similarity.cell(row=7, column=4).value = "=IF(C7=\"\",\"\",IF(ISNA(MATCH(C7,INDIRECT(\"'Word Vectors \"&B3&\"'!A:A\"),0)),\"Word not found\",MATCH(C7,INDIRECT(\"'Word Vectors \"&B3&\"'!A:A\"),0)-1))"
    formula_template = "=IF(OR(A7=\"\",C7=\"\",NOT(ISNUMBER(B7)),NOT(ISNUMBER(D7))),\"\",SUMPRODUCT(INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(B7+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(B7+1)),INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(D7+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(D7+1)))/(SQRT(SUMPRODUCT(INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(B7+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(B7+1)),INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(B7+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(B7+1))))*SQRT(SUMPRODUCT(INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(D7+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(D7+1)),INDIRECT(\"'Word Vectors \"&B3&\"'!B\"&(D7+1)&\":\"&IF(B3=\"Twitter\",\"Z\",\"AZ\")&(D7+1))))))"
    ws_similarity.cell(row=7, column=5).value = formula_template
    
    ws_similarity.cell(row=8, column=1).value = "Tip:"
    ws_similarity.cell(row=8, column=1).font = Font(bold=True, italic=True)
    tip_text = "You can copy row 7 down to compare multiple word pairs. Change the dataset in B3 to switch between Twitter and Wiki datasets."
    ws_similarity.cell(row=8, column=2).value = tip_text
    ws_similarity.merge_cells('B8:E8')
    ws_similarity.cell(row=8, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws_similarity.cell(row=8, column=2).font = Font(italic=True)
    
    ws_similarity.column_dimensions['A'].width = 15
    ws_similarity.column_dimensions['B'].width = 20
    ws_similarity.column_dimensions['C'].width = 15
    ws_similarity.column_dimensions['D'].width = 20
    ws_similarity.column_dimensions['E'].width = 25
    
    # Now add formulas to Search sheet with dropdown
    print("Adding formulas to Search sheet...")
    ws_search = wb['Search']
    
    # Define highlight color for input cells
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws_search.cell(row=1, column=1).value = "TOP 10 SIMILAR WORDS FINDER"
    ws_search.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_search.merge_cells('A1:C1')
    ws_search.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Add dataset selector dropdown
    ws_search.cell(row=2, column=1).value = "Dataset:"
    ws_search.cell(row=2, column=1).font = Font(bold=True)
    ws_search.cell(row=2, column=2).value = "Twitter"  # Default
    ws_search.cell(row=2, column=2).fill = input_fill  # Highlight input cell
    
    dv_search = DataValidation(type="list", formula1='"Twitter,Wiki"', allow_blank=False)
    ws_search.add_data_validation(dv_search)
    dv_search.add(ws_search.cell(row=2, column=2))
    
    ws_search.cell(row=3, column=1).value = "Enter a word:"
    ws_search.cell(row=3, column=1).font = Font(bold=True)
    ws_search.cell(row=3, column=2).value = "people"  # Example word - exists in both datasets
    ws_search.cell(row=3, column=2).fill = input_fill  # Highlight input cell
    
    ws_search.cell(row=4, column=1).value = "Instructions:"
    ws_search.cell(row=4, column=1).font = Font(bold=True)
    top_instructions = "Select a dataset from the dropdown (B2), then enter any word from that dataset in cell B3. The sheet will automatically find the top 10 most similar words based on cosine similarity."
    ws_search.cell(row=4, column=2).value = top_instructions
    ws_search.merge_cells('B4:C4')
    ws_search.cell(row=4, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    
    headers_top = ['Rank', 'Word', 'Similarity']
    for col_idx, header in enumerate(headers_top, start=1):
        cell = ws_search.cell(row=6, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Determine which dataset to use based on dropdown
    num_words_twitter = len(df_twitter)
    num_words_wiki = len(df_wiki)
    
    # Create helper columns for Twitter similarities
    print(f"  Creating {num_words_twitter} helper columns for Twitter similarities...")
    for i in range(num_words_twitter):
        col_letter = get_column_letter(4 + i)
        word_row = i + 2
        similarity_formula = f"""=IF(OR($B$2<>"Twitter",$B$3="",ISNA(MATCH($B$3,'Word Vectors Twitter'!A:A,0)),'Word Vectors Twitter'!A{word_row}=$B$3),"",SUMPRODUCT(INDIRECT("'Word Vectors Twitter'!B"&MATCH($B$3,'Word Vectors Twitter'!A:A,0)&":Z"&MATCH($B$3,'Word Vectors Twitter'!A:A,0)),INDIRECT("'Word Vectors Twitter'!B{word_row}:Z{word_row}"))/(SQRT(SUMPRODUCT(INDIRECT("'Word Vectors Twitter'!B"&MATCH($B$3,'Word Vectors Twitter'!A:A,0)&":Z"&MATCH($B$3,'Word Vectors Twitter'!A:A,0)),INDIRECT("'Word Vectors Twitter'!B"&MATCH($B$3,'Word Vectors Twitter'!A:A,0)&":Z"&MATCH($B$3,'Word Vectors Twitter'!A:A,0))))*SQRT(SUMPRODUCT(INDIRECT("'Word Vectors Twitter'!B{word_row}:Z{word_row}"),INDIRECT("'Word Vectors Twitter'!B{word_row}:Z{word_row}")))))"""
        ws_search.cell(row=3, column=4+i).value = similarity_formula
    
    # Create helper columns for Wiki similarities (starting after Twitter columns)
    print(f"  Creating {num_words_wiki} helper columns for Wiki similarities...")
    start_wiki_col = 4 + num_words_twitter
    for i in range(num_words_wiki):
        col_letter = get_column_letter(start_wiki_col + i)
        word_row = i + 2
        similarity_formula = f"""=IF(OR($B$2<>"Wiki",$B$3="",ISNA(MATCH($B$3,'Word Vectors Wiki'!A:A,0)),'Word Vectors Wiki'!A{word_row}=$B$3),"",SUMPRODUCT(INDIRECT("'Word Vectors Wiki'!B"&MATCH($B$3,'Word Vectors Wiki'!A:A,0)&":AZ"&MATCH($B$3,'Word Vectors Wiki'!A:A,0)),INDIRECT("'Word Vectors Wiki'!B{word_row}:AZ{word_row}"))/(SQRT(SUMPRODUCT(INDIRECT("'Word Vectors Wiki'!B"&MATCH($B$3,'Word Vectors Wiki'!A:A,0)&":AZ"&MATCH($B$3,'Word Vectors Wiki'!A:A,0)),INDIRECT("'Word Vectors Wiki'!B"&MATCH($B$3,'Word Vectors Wiki'!A:A,0)&":AZ"&MATCH($B$3,'Word Vectors Wiki'!A:A,0))))*SQRT(SUMPRODUCT(INDIRECT("'Word Vectors Wiki'!B{word_row}:AZ{word_row}"),INDIRECT("'Word Vectors Wiki'!B{word_row}:AZ{word_row}")))))"""
        ws_search.cell(row=3, column=start_wiki_col+i).value = similarity_formula
    
    # Create formulas for top 10 matches
    last_twitter_col = get_column_letter(4 + num_words_twitter - 1)
    last_wiki_col = get_column_letter(start_wiki_col + num_words_wiki - 1)
    
    for rank in range(1, 11):
        row_num = 6 + rank
        
        ws_search.cell(row=row_num, column=1).value = rank
        
        # Similarity value - use LARGE on the appropriate range based on dataset
        similarity_formula = f"""=IF(OR($B$2="",$B$3=""),"",IF($B$2="Twitter",IF(ISNA(MATCH($B$3,'Word Vectors Twitter'!A:A,0)),"",LARGE(D3:{last_twitter_col}3,{rank})),IF(ISNA(MATCH($B$3,'Word Vectors Wiki'!A:A,0)),"",LARGE({get_column_letter(start_wiki_col)}3:{last_wiki_col}3,{rank}))))"""
        ws_search.cell(row=row_num, column=3).value = similarity_formula
        
        # Word - find which word has this similarity value
        word_formula = f"""=IF(OR($B$2="",$B$3="",C{row_num}=""),"",IF($B$2="Twitter",IF(ISNA(MATCH($B$3,'Word Vectors Twitter'!A:A,0)),"",INDEX('Word Vectors Twitter'!A:A,MATCH(C{row_num},D3:{last_twitter_col}3,0)+1)),IF(ISNA(MATCH($B$3,'Word Vectors Wiki'!A:A,0)),"",INDEX('Word Vectors Wiki'!A:A,MATCH(C{row_num},{get_column_letter(start_wiki_col)}3:{last_wiki_col}3,0)+1))))"""
        ws_search.cell(row=row_num, column=2).value = word_formula
    
    note_row = 6 + 11
    ws_search.cell(row=note_row, column=1).value = "Note:"
    ws_search.cell(row=note_row, column=1).font = Font(bold=True, italic=True)
    note_text = "Similarity scores range from -1 to 1. Higher values indicate more similar words. Select dataset from dropdown in B2, then enter word in B3."
    ws_search.cell(row=note_row, column=2).value = note_text
    ws_search.merge_cells(f'B{note_row}:C{note_row}')
    ws_search.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws_search.cell(row=note_row, column=2).font = Font(italic=True)
    
    # Hide all helper columns
    for i in range(num_words_twitter + num_words_wiki):
        col_letter = get_column_letter(4 + i)
        ws_search.column_dimensions[col_letter].hidden = True
    
    ws_search.column_dimensions['A'].width = 10
    ws_search.column_dimensions['B'].width = 20
    ws_search.column_dimensions['C'].width = 15
    
    # Update Instructions sheet to format formulas with quotes
    ws_instructions = wb['Instructions']
    # The formulas are already quoted in the DataFrame, but let's make sure they're displayed correctly
    
    wb.save(output_file)
    print(f"✓ Excel file created: {output_file}")
    print(f"  - Sheet 'Word Vectors Twitter': Contains Twitter word vectors (25 dims)")
    print(f"  - Sheet 'Word Vectors Wiki': Contains Wiki word vectors (50 dims)")
    print(f"  - Sheet 'Search': Enter a word in B3, select dataset in B2 dropdown")
    print(f"  - Sheet 'Cosine Similarity': Select dataset in B3 dropdown")
    print(f"  - Sheet 'Instructions': Contains formula documentation")
    print(f"  - Input cells are highlighted in yellow")

if __name__ == "__main__":
    create_excel_with_formulas()
